import tkinter as tk
import openpyxl as xl
from openpyxl import Workbook
import sqlite3

# Establish connection to SQLite3 database
conn = sqlite3.connect('spiers_database.db')

# Enable foreign key support
conn.execute('PRAGMA foreign_keys = ON;')

# Link cursor
cursor = conn.cursor()

# Report Frame
class ReportFrame(tk.Frame):
    def __init__(self, master, controller):
        tk.Frame.__init__(self, master)

        self.workbook_title = "Request Report.xlsx"

        self.request_id = 1

        self.report_timestamp = tk.StringVar()
        self.report_timestamp.set("4/13/2025, 12:00")

        self.report = tk.StringVar()
        self.report.set("Dummy Report")

        try:
            self.wb = xl.load_workbook(self.workbook_title)  # creates a workbook object, argument is the name of workbook
        except FileNotFoundError: ## if workbook doesn't exist, create new workbook, save with intended title, and load it again
            self.wb = Workbook()
            self.wb.save (self.workbook_title)
            self.wb = xl.load_workbook(self.workbook_title)

        report_sheet = self.wb.active
        report_sheet.title = "Reports"

        timestamp_title_cell = report_sheet["A1"]
        request_title_cell = report_sheet["B1"]
        report_title_cell = report_sheet["C1"]

        if (report_sheet["A1"].value != "Report Timestamp" ## if title columns have wrong title, give them the right title
         or report_sheet["B1"].value != "Request ID"
         or report_sheet["C1"].value != "Report"):
            timestamp_title_cell.value = "Report Timestamp"
            request_title_cell.value = "Request ID"
            report_title_cell.value = "Report"

        timestamp_cell = report_sheet[f"A{self.request_id + 1}"]  # increment so cell titles are not over-written
        request_cell = report_sheet[f"B{self.request_id + 1}"]
        report_cell = report_sheet[f"C{self.request_id + 1}"]

        timestamp_cell.value = self.report_timestamp.get() ## Assign values to cells
        request_cell.value = self.request_id
        report_cell.value = self.report.get()

        for col in report_sheet.columns: # Automatically Adjust Cell Width
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for row in range(1, report_sheet.max_row + 1):
                cell = report_sheet[f"{column}{row}"]
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            report_sheet.column_dimensions[column].width = adjusted_width

        self.report_label = tk.Label(master = self)
        self.report_label.config(text = "Please Make Any Final Edits To The Report")
        self.report_label.grid(row = 0, column = 1, padx = 10, pady = 10)

        self.report_text = tk.Text(master = self)
        self.report_text.config ()
        self.report_label.grid(row = 1, column = 1, padx = 10, pady = 10)

        self.forward_button = tk.Button(master = self)
        self.forward_button.config(width = 20, text ="Confirm", command = lambda : close_app())
        self.forward_button.grid(row = 2, column = 1, padx = 10, pady = 10)

        self.back_button = tk.Button(master=self)
        self.back_button.config(width=20, text="Back", command=lambda: controller.back_button())
        self.back_button.grid(row=2, column=0, padx=10, pady=10, sticky = "SW")

        def close_app():
            self.wb.save(self.workbook_title)
            controller.destroy()

conn.commit()
conn.close()