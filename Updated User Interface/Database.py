import mariadb
import sys
from PIL import Image
import io
import openpyxl as xl
from openpyxl import Workbook

# TODO: Add Mock Pictures

# create a class that stores the data
class DatabaseManager:
    # the username of the database on AWS
    rds_db_user = "spiersGroup1"

    # the password of the database on AWS
    rds_db_password = "spiersGroup1DB"

    # Replace with your AWS RDS endpoint
    rds_db_host = "spiersgroup1.cxg20au4crax.us-east-2.rds.amazonaws.com"

    # Default port for MariaDB
    rds_db_port = 3306

    # the name of the database on AWS
    rds_db_name = "spiersGroup1DB"

    # connector and cursor variables to traverse the database
    rds_conn = None
    rds_cursor = None

    # Local MariaDB Database Connection Parameters
    local_db_user = "root"
    local_db_password = ""  # Replace with system MariaDB password
    local_db_host = "localhost"  # Assuming local MariaDB is on your machine
    local_db_port = 3306  # Default port for MariaDB
    local_db_name = "mariadb_data"
    local_conn = None
    local_cursor = None


    @classmethod
    def establish_connection(cls, db_user, db_password, db_host, db_port, db_name):
        # try to connect to AWS database
        try:
            conn = mariadb.connect(
                user=db_user,
                password=db_password,
                host=db_host,
                port=db_port,
                database=db_name)

            # Disable Auto-Commit so we only commit when changes are ready
            conn.autocommit = False

            # return connector so it can be used in other areas of code
            return conn
        except mariadb.Error as e:
            print(f"Error connecting to MariaDB Platform: {e}")
            sys.exit(1)

    @classmethod
    # Local Database setup
    def setup_mock_database(cls, db_conn):
        try:
            # try to connect to database
            conn = db_conn

            # get a cursor to carry out operations on the database
            cursor = conn.cursor()

            # create tables with SQL commands that have the listed parameters
            create_tables_sql = [
                """
                CREATE TABLE IF NOT EXISTS battery_state (
                    state_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    state_desc TEXT
                );
                """, # this table stores the ba
                """
                CREATE TABLE IF NOT EXISTS client_status (
                    client_status_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    client_status_desc TEXT
                );
                """,
                """
                CREATE TABLE if not exists clients(
                    client_id integer PRIMARY KEY AUTO_INCREMENT,
                    client_desc TEXT,
                    client_status_id integer,
                    FOREIGN KEY (client_status_id) REFERENCES client_status (client_status_id) ON DELETE CASCADE
                );
                """,
                """
                CREATE TABLE IF NOT EXISTS works (
                    work_type_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    work_type_name TEXT,
                    parent_work_type_id INT,
                    FOREIGN KEY (parent_work_type_id) REFERENCES works (work_type_id) ON DELETE CASCADE
                );
                """,
                """
                CREATE TABLE IF NOT EXISTS employees (
                    user_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    first_name TEXT,
                    last_name TEXT,
                    password TEXT,
                    is_admin INTEGER
                );
                """,
                """
                CREATE TABLE IF NOT EXISTS locations (
                    location_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    location_shorthand TEXT,
                    location_description TEXT
                );
                """,
                """
                CREATE TABLE IF NOT EXISTS batteries (
                    serial_number VARCHAR(100) PRIMARY KEY,
                    part_number INTEGER,
                    item_type INTEGER,
                    part_description TEXT,
                    location INTEGER,
                    picture LONGBLOB,
                    FOREIGN KEY (location) REFERENCES locations(location_id) ON DELETE CASCADE
                );
                """,
                """
                CREATE TABLE IF NOT EXISTS requests (
                    request_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    request_time TIMESTAMP UNIQUE,
                    serial_number VARCHAR(100),
                    work_type_id INTEGER,
                    user_id INTEGER,
                    state_id INTEGER,
                    client_id INTEGER,
                    FOREIGN KEY (serial_number) REFERENCES batteries(serial_number) ON DELETE CASCADE,
                    FOREIGN KEY (work_type_id) REFERENCES works(work_type_id) ON DELETE CASCADE,
                    FOREIGN KEY (user_id) REFERENCES employees(user_id) ON DELETE CASCADE,
                    FOREIGN KEY (state_id) REFERENCES battery_state(state_id) ON DELETE CASCADE,
                    FOREIGN KEY (client_id) REFERENCES clients (client_id) ON DELETE CASCADE
                );
                """,
                """
                CREATE TABLE IF NOT EXISTS reports (
                    request_id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    request_time TIMESTAMP UNIQUE,
                    report_desc TEXT,
                    FOREIGN KEY (request_id) REFERENCES requests(request_id) ON DELETE CASCADE,
                    FOREIGN KEY (request_time) REFERENCES requests(request_time) ON DELETE CASCADE
                );
                """
            ]
            for statement in create_tables_sql:
                # Execute each SQL statement to create the tables
                cursor.execute(statement)
                # save the data changes
                cls.save_changes(conn)

        # if there is an error, then print the error, but still run the program
        except mariadb.Error as e:
            print(f"Error: {e}")

    @classmethod
    def add_mock_data (cls, db_conn):
        try:
            # Establish connection to Mariadb database
            conn = db_conn

            # Link cursor
            cursor = conn.cursor()

            """
            Create a list of tuples containing the state_id and the battery state. 
            IDs are specified so data doesn't duplicate; thus, we can save storage costs. 
            """
            battery_state = [
                (1, "New"),
                (2, "Old"),
                (3, "Death Row")
            ]

            # Insert each battery state into the battery state table
            for state_id, state_desc in battery_state:
                cursor.execute("""
                    INSERT INTO battery_state (state_id, state_desc)
                    VALUES (?, ?)
                """, (state_id, state_desc))

            cls.save_changes(conn)

            client_status = [
                (1, "Suppliers"),
                (2, "Customers")
            ]

            for client_status_id, client_status_desc in client_status:
                cursor.execute("""
                    INSERT INTO client_status (client_status_id, client_status_desc)
                    VALUES (?, ?)
                """, (client_status_id, client_status_desc))

            cls.save_changes(conn)

            clients = [
                (1, None, 1),
                (2, "Battery New", 1),
                (3, "Loyal", 2),
                (4, "High Income", 2),
                (5, "Thrifty", 2),
                (6, "Old Reliable", 1),
                (7, "Never Death Row", 1)
            ]

            # Insert each client into the client table
            for client_id, client_desc, client_status_id in clients:
                cursor.execute("""
                               INSERT INTO clients (client_id, client_desc, client_status_id)
                               VALUES (?, ?, ?)
                               """, (client_id, client_desc, client_status_id))

            cls.save_changes(conn)

            works = [
                (1, "Find", None),
                (2, "Receive", 1),
                (3, "Ship", 1),
                (4, "Move", 1),
                (5, "Update Battery Status", 2),
                (6, "New Battery Work", 5),
                (7, "Old Battery Work", 5),
                (8, "Death Row Battery Work", 5),
                (9, "Store", 6),
                (10, "Diagnostic Analysis", 7),
                (11, "Disassembly", 7),
                (12, "Repair", 7),
                (13, "Re-assembly", 7),
                (14, "Testing", 7),
                (15, "Re-certification", 7),
                (16, "Take Apart", 8),
                (17, "Shred (pieces)", 8),
                (18, "Shred (powder)", 8),
                (19, "Make new battery", 8),
                (20, "Take Picture", None),
                (21, "Intake New Item", None)
            ]

            for work_type_id, work_type_name, parent_work_type_id in works:
                cursor.execute("""
                            INSERT INTO works (work_type_id, work_type_name, parent_work_type_id)
                            VALUES (?, ?, ?)
                        """, (work_type_id, work_type_name, parent_work_type_id))

            cls.save_changes(conn)

            # Insert each employee into the employees table
            employees = [
                (1, "John", "Smith", "123password", 1),
                (2, "Emily", "Johnson", "p123assword", 0),
                (3, "Michael", "Williams", "pa123ssword", 0),
                (4, "Sarah", "Brown", "pas123sword", 1),
                (5, "James", "Davis", "pass123word", 0),
                (6, "Jessica", "Wilson", "passw123ord", 0),
                (7, "David", "Moore", "passwo123rd", 1),
                (8, "Olivia", "Taylor", "passwor123d", 0),
                (9, "Daniel", "Anderson", "password123", 0)
            ]

            for user_id, first_name, last_name, password, is_admin in employees:
                cursor.execute('''
                    INSERT INTO employees (user_id, first_name, last_name, password, is_admin)
                    VALUES (?, ?, ?, ?, ?)
                ''', (user_id, first_name, last_name, password, is_admin))

            cls.save_changes(conn)

            # Insert each location into the location table
            locations = [
                ("FS1", "floor space 1"),
                ("FS2", "floor space 2"),
                ("FS3", "floor space 3"),
                ("SS1", "shelf space 1"),
                ("SS2", "shelf space 2"),
                ("SS3", "shelf space 3")
            ]

            for location_shorthand, location_description in locations:
                cursor.execute("""
                    INSERT INTO locations (location_shorthand, location_description)
                    VALUES (?, ?)
                """, (location_shorthand, location_description))

            cls.save_changes(conn)

            battery_1_image = Image.open("photo_20250423_121958.jpg")
            og_width, og_height = battery_1_image.size
            battery_1_image = battery_1_image.resize((og_width // 3, og_height // 3), Image.LANCZOS)

            # Convert the image to binary data (BLOB format)
            with io.BytesIO() as byte_io:
                battery_1_image.save(byte_io, format="JPEG")
                battery_1_image_data = byte_io.getvalue()

            battery_2_image = Image.open("20250425_001602.jpg")
            og_width, og_height = battery_2_image.size
            battery_2_image = battery_2_image.resize((og_width // 3, og_height // 3), Image.LANCZOS)

            # Convert the image to binary data (BLOB format)
            with io.BytesIO() as byte_io:
                battery_2_image.save(byte_io, format="JPEG")
                battery_2_image_data = byte_io.getvalue()

            battery_3_image = Image.open("20250425_001614.jpg")
            og_width, og_height = battery_3_image.size
            battery_3_image = battery_3_image.resize((og_width // 3, og_height // 3), Image.LANCZOS)

            # Convert the image to binary data (BLOB format)
            with io.BytesIO() as byte_io:
                battery_3_image.save(byte_io, format="JPEG")
                battery_3_image_data = byte_io.getvalue()

            batteries = [
                ("battery_1_serial", 1, 1, "battery 1", 1, battery_1_image_data),
                ("battery_2_serial", 1, 1, "battery 2", 2, battery_2_image_data),
                ("battery_3_serial", 2, 1, "battery 3", 3, battery_3_image_data),
                ("battery_4_serial", 2, 1, "battery 4", 1, battery_1_image_data),
                ("battery_5_serial", 2, 1, "battery 5", 2, battery_2_image_data),
                ("battery_6_serial", 2, 1, "battery 6", 3, battery_3_image_data)
            ]

            for serial_number, part_number, item_type, part_description, location, picture in batteries:
                cursor.execute("""
                    INSERT INTO batteries (serial_number, part_number, item_type, part_description, location, picture)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (serial_number, part_number, item_type, part_description, location, picture))

            cls.save_changes(conn)

        except mariadb.Error as e:
           print(f"Error: {e}")

    @classmethod
    def setup_local_database(cls):
        cls.local_conn =  cls.establish_connection(cls.local_db_user, cls.local_db_password, cls.local_db_host, cls.local_db_port, cls.local_db_name)
        cls.local_cursor = cls.local_conn.cursor()
        cls.setup_mock_database(cls.local_conn)
        cls.add_mock_data(cls.local_conn)

    """
    @classmethod
    def get_local_conn (cls):
        return cls.local_conn
    """

    @classmethod
    def setup_rds_database(cls):
        cls.rds_conn = cls.establish_connection(cls.rds_db_user, cls.rds_db_password, cls.rds_db_host, cls.rds_db_port, cls.rds_db_name)
        cls.rds_cursor = cls.rds_conn.cursor()
        cls.setup_mock_database(cls.rds_conn)
        cls.add_mock_data(cls.rds_conn)

    @classmethod
    def clear_everything(cls):
        commands = [
            """drop table if exists reports;""",
            """drop table if exists requests;""",
            """drop table if exists batteries;""",
            """drop table if exists locations;""",
            """drop table if exists employees;""",
            """drop table if exists works;""",
            """drop table if exists battery_state;""",
            """drop table if exists suppliers;""",
            """drop table if exists customers;""",
            """drop table if exists clients;""",
            """drop table if exists client_status;""",
            """drop table if exists warehouses;"""
        ]

        cls.rds_conn = cls.establish_connection(cls.rds_db_user, cls.rds_db_password, cls.rds_db_host, cls.rds_db_port,
                                            cls.rds_db_name)
        cls.rds_cursor = cls.rds_conn.cursor()

        for instruction in commands:
            cls.rds_cursor.execute(instruction)

        cls.rds_conn.commit()
        cls.rds_conn.close()

        cls.local_conn = cls.establish_connection(cls.local_db_user, cls.local_db_password, cls.local_db_host, cls.local_db_port,
                                                cls.local_db_name)
        cls.local_cursor = cls.local_conn.cursor()

        for instruction in commands:
            cls.local_cursor.execute(instruction)

        cls.local_conn.commit()
        cls.local_conn.close()

        # set title of excel file to save reports in
        workbook_title = "Request Report.xlsx"

        # creates a workbook object, argument is the name of workbook
        try:
            wb = xl.load_workbook(workbook_title)
        ## if workbook doesn't exist, create new workbook, save with intended title, and load it again
        except FileNotFoundError:
            wb = Workbook()
            wb.save(workbook_title)
            wb = xl.load_workbook(workbook_title)

        # choose the report sheet in excel file
        report_sheet = wb["Reports"]

        # Clear the report sheet's contents
        for row in report_sheet.iter_rows():
            for cell in row:
                cell.value = None

        # save the changes
        wb.save(workbook_title)

    @classmethod
    def get_rds_conn (cls):
        return cls.rds_conn

    @classmethod
    def save_changes(cls, db_conn):
        db_conn.commit()

    @classmethod
    def close_connection(cls, db_conn):
        db_conn.close()