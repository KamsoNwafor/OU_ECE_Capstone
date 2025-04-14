import tkinter as tk
import openpyxl as xl
from openpyxl import Workbook

class App (tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        # creating a container
        container = tk.Frame(self)
        container.pack(fill = tk.BOTH, expand = True)

        # initializing frames to an empty array
        self.frames = {}

        # iterating through a tuple consisting of the different page layouts
        for F in (StartPage, WarehousePage, ReportPage):
            frame = F(container, self)

            # initializing frame of that object from start page, warehouse page, report page respectively with for loop
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_page(StartPage)

    # to display the current frame passed as a parameter
    def show_page(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

class StartPage(tk.Frame):
    def __init__(self, master, controller):
        tk.Frame.__init__(self, master)

        start_label = tk.Label(master = self, text="Spiers New Technologies")
        start_label.grid(row = 0, column = 1, padx = 10, pady = 10)

        start_button = tk.Button(master = self, width=25, text="Start", command = lambda : controller.show_page(WarehousePage)) ## set button width to be 25% of screen width
        start_button.grid(row = 1, column = 1, padx = 10, pady = 10)

class WarehousePage(tk.Frame):
    def __init__(self, master, controller):
        tk.Frame.__init__(self, master)

        self.warehouse_label = tk.Label(master=self, text="Which Warehouse Is This?")
        self.warehouse_label.grid (row = 0, column = 1, padx = 10, pady = 10)

        self.report_button = tk.Button(master = self)
        self.report_button.config(width = 20, text = "Edit Report", command = lambda : controller.show_page(ReportPage))
        self.report_button.grid (row = 3, column = 2, padx = 10, pady = 10, sticky="SE")

        self.warehouses = ("Warehouse 1", "Warehouse 2", "Warehouse 3")

        self.location = tk.StringVar()
        self.location.set("")

        self.warehouse_input = tk.Entry(master=self)
        self.warehouse_input.config(textvariable=self.location)
        self.warehouse_input.grid(row=1, column=1, padx=10, pady=10)
        self.warehouse_input.bind('<KeyRelease>', self.checkkey)

        self.warehouse_list = tk.Listbox(master=self)
        self.warehouse_list.config()
        self.warehouse_list.grid(row=2, column=1, padx=10, pady=10)
        self.list_update(self.warehouses)

        self.back_button = tk.Button(master=self)
        self.back_button.config(width=20, text="Back", command=lambda: controller.show_page(StartPage))
        self.back_button.grid(row = 3, column = 0, padx = 10, pady = 10, sticky="SW")

    def checkkey(self, event):

        value = event.widget.get()
        print(value)

        # get data from l
        if value == '':
            data = self.warehouses
        else:
            data = []
            for item in self.warehouses:
                if value.lower() in item.lower():
                    data.append(item)

                    # update data in listbox
        self.list_update(data)

    def list_update(self, data):

        # clear previous data
        self.warehouse_list.delete(0, 'end')

        # put new data
        for item in data:
            self.warehouse_list.insert('end', item)

            ## Report Frame
class ReportPage(tk.Frame):
    def __init__(self, master, controller):
        tk.Frame.__init__(self, master)

        self.workbook_title = "Request Report.xlsx"

        self.request_id = 1

        self.report_timestamp = tk.StringVar()
        self.report_timestamp.set("4/13/2025, 12:00")

        self.report = tk.StringVar()
        self.report.set("Dummy Report")

        try:
            self.wb = xl.load_workbook(self.workbook_title)  # creates a workbook object, argument is the name of workbook
        except FileNotFoundError: ## if workbook doesn't exist, create new workbook, save with intended title, and load it again
            self.wb = Workbook()
            self.wb.save (self.workbook_title)
            self.wb = xl.load_workbook(self.workbook_title)

        report_sheet = self.wb.active
        report_sheet.title = "Reports"

        timestamp_title_cell = report_sheet["A1"]
        request_title_cell = report_sheet["B1"]
        report_title_cell = report_sheet["C1"]

        if (report_sheet["A1"].value != "Report Timestamp" ## if title columns have wrong title, give them the right title
         or report_sheet["B1"].value != "Request ID"
         or report_sheet["C1"].value != "Report"):
            timestamp_title_cell.value = "Report Timestamp"
            request_title_cell.value = "Request ID"
            report_title_cell.value = "Report"

        timestamp_cell = report_sheet[f"A{self.request_id + 1}"]  # increment so cell titles are not over-written
        request_cell = report_sheet[f"B{self.request_id + 1}"]
        report_cell = report_sheet[f"C{self.request_id + 1}"]

        timestamp_cell.value = self.report_timestamp.get() ## Assign values to cells
        request_cell.value = self.request_id
        report_cell.value = self.report.get()

        for col in report_sheet.columns: # Automatically Adjust Cell Width
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for row in range(1, report_sheet.max_row + 1):
                cell = report_sheet[f"{column}{row}"]
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            report_sheet.column_dimensions[column].width = adjusted_width

        report_label = tk.Label(master = self)
        report_label.config(text = "Please Make Any Final Edits To The Report")
        report_label.grid(row = 0, column = 4, padx = 10, pady = 10)

        report_text = tk.Text(master = self)
        report_text.config ()
        report_label.grid(row = 1, column = 1, padx = 10, pady = 10)

        report_button = tk.Button(master = self)
        report_button.config(width = 25, text = "Submit Report", command = lambda : self.close_app())
        report_button.grid(row = 2, column = 1, padx = 10, pady = 10)

    def close_app(self):
        self.wb.save(self.workbook_title)
        self.destroy()

root = App()
root.title("Spiers New Technologies")
root.geometry("800x480")

root.mainloop()