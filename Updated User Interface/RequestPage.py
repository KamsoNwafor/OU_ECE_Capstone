import io
import os
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from Database import DatabaseManager as dbm
from datetime import datetime
from PIL import Image

class RequestFrame(tk.Frame):
    frame_index = 13

    def __init__(self, master, controller):
        tk.Frame.__init__(self, master, bg="#fafafa")
        self.controller = controller

        # Database Connection
        self.rds_conn = dbm.get_rds_conn()
        self.rds_cursor = self.rds_conn.cursor()

        # Initialize Variables
        self.reset_report()

        # Picture resize settings
        self.max_width = 800 // 3
        self.max_height = 480 // 3

        # Excel Workbook Setup
        self.workbook_title = "Request Report.xlsx"
        try:
            self.wb = xl.load_workbook(self.workbook_title)
        except FileNotFoundError:
            self.wb = Workbook()
            self.wb.save(self.workbook_title)
            self.wb = xl.load_workbook(self.workbook_title)

        self.report_sheet = self.wb.active
        self.report_sheet.title = "Reports"
        self.setup_report_sheet_headers()

        # === UI Layout ===

        # Header
        header = tk.Frame(self, bg="#4CAF50")
        header.pack(fill="x")
        tk.Label(header, text="Step 6: Final Report", font=("Roboto", 14, "bold"), bg="#4CAF50", fg="#FFFFFF").pack(pady=10)

        # Content
        content = tk.Frame(self, bg="#f0f0f0", bd=1, relief="solid")
        content.pack(padx=10, pady=10, fill="both", expand=True)
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(1, weight=1)

        self.report_label = tk.Label(content, text="Please Make Any Final Edits To The Report",
                                     font=("Roboto", 12, "bold"), bg="#f0f0f0", fg="#212121")
        self.report_label.grid(row=0, column=0, pady=(10, 5))

        # Text Area (Reduced Height)
        self.report_text = tk.Text(content, height= 8, wrap=tk.WORD, font=("Roboto", 11))
        self.report_text.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        # Navigation Buttons
        nav_frame = tk.Frame(content, bg="#f0f0f0")
        nav_frame.grid(row=2, column=0, pady=10)
        nav_frame.grid_columnconfigure(0, weight=1)
        nav_frame.grid_columnconfigure(1, weight=1)

        self.back_button = ttk.Button(nav_frame, text="Back", style="Secondary.TButton", command=self.previous_page)
        self.back_button.grid(row=0, column=0, padx=5)

        self.submit_button = ttk.Button(nav_frame, text="Confirm", style="Primary.TButton", command=self.complete_report)
        self.submit_button.grid(row=0, column=1, padx=5)

    def setup_report_sheet_headers(self):
        """Create headers for the Excel sheet if not present."""
        self.request_title_cell = self.report_sheet["A1"]
        self.timestamp_title_cell = self.report_sheet["B1"]
        self.report_title_cell = self.report_sheet["C1"]

        if (self.request_title_cell.value != "Request ID" or
            self.timestamp_title_cell.value != "Report Timestamp" or
            self.report_title_cell.value != "Report"):
            self.request_title_cell.value = "Request ID"
            self.timestamp_title_cell.value = "Report Timestamp"
            self.report_title_cell.value = "Report"

    def complete_report(self):
        """Submit final report."""
        self.submit_request()
        self.submit_report()
        self.wb.save(self.workbook_title)
        messagebox.showinfo("Success", "Report Submitted Successfully!")
        self.controller.destroy()

    def previous_page(self):
        """Go back to the previous page and clear the report."""
        self.controller.back_button()
        self.report_text.delete("1.0", tk.END)
        self.reset_report()

    def load_report(self):
        """Load all user input data to create the report."""
        self.rds_cursor.execute("SELECT request_id FROM requests ORDER BY request_id DESC LIMIT 1;")
        result = self.rds_cursor.fetchall()
        self.request_id = 1 if not result else result[0][0] + 1

        self.curr_time = datetime.now()
        self.request_timestamp.set(self.curr_time.strftime("%Y-%m-%d %H:%M:%S"))

        # Fetch data
        self.serial_num = self.controller.selected_battery_serial_number
        self.work_type_id = self.controller.selected_task_id
        self.user_id = self.controller.selected_user_id
        self.emotion = self.controller.selected_emotion
        self.adjective = self.controller.selected_adjective
        self.state_id = self.controller.selected_state_id or "1"
        self.client_id = self.controller.selected_client_id or "1"
        self.actions = self.controller.selected_actions or ["1"]

        if self.controller.selected_picture:
            self.picture = self.controller.selected_picture.resize((self.max_width, self.max_height), Image.Resampling.LANCZOS)
            with io.BytesIO() as byte_io:
                self.picture.save(byte_io, format="JPEG")
                self.picture_data = byte_io.getvalue()

        self.new_location = self.fetch_location_desc(self.controller.selected_location_id)
        self.old_location = self.fetch_location_desc(self.controller.old_location_id) if self.controller.old_location_id else "No Location"

        if self.controller.input_battery_desc:
            self.battery_desc = self.controller.input_battery_desc
        else:
            self.rds_cursor.execute("SELECT part_description FROM batteries WHERE serial_number = %s", (self.serial_num,))
            result = self.rds_cursor.fetchall()
            self.battery_desc = result[0][0] if result else "Unknown Part"

        self.part_num = self.controller.selected_part_number
        self.item_type = self.controller.selected_item_type

        self.rds_cursor.execute("SELECT first_name, last_name FROM employees WHERE user_id = %s", (self.user_id,))
        result = self.rds_cursor.fetchall()
        self.employee = f"{result[0][0]} {result[0][1]}" if result else "Unknown Employee"

        self.rds_cursor.execute("SELECT client_desc, client_status_id FROM clients WHERE client_id = %s", (self.client_id,))
        result = self.rds_cursor.fetchall()
        self.client = result[0][0] if result else "Unknown Client"
        self.client_status = self.fetch_client_status_desc(result[0][1]) if result else "Unknown Status"

        self.rds_cursor.execute("SELECT state_desc FROM battery_state WHERE state_id = %s", (self.state_id,))
        result = self.rds_cursor.fetchall()
        self.battery_state = result[0][0] if result else "Unknown State"

        self.battery_actions = ""
        for action in self.actions:
            self.rds_cursor.execute("SELECT work_type_name FROM works WHERE work_type_id = %s", (action,))
            result = self.rds_cursor.fetchall()
            if result:
                self.battery_actions += f"{result[0][0]}, "

        final_report = self.compose_report()
        self.report.set(final_report)
        self.report_text.delete("1.0", tk.END)
        self.report_text.insert("end", self.report.get())

    def compose_report(self):
        """Compose the text that will appear in the report."""
        if self.controller.input_battery_desc:
            return (f"{self.employee} added {self.battery_desc} to the database.\n"
                    f"Serial Number: {self.serial_num}\n"
                    f"Part Number: {self.part_num}\n"
                    f"Item Type: {self.item_type}\n"
                    f"Location: {self.new_location}\n"
                    f"Now {self.employee} is {self.emotion}, feeling the task was {self.adjective}.")
        else:
            if self.work_type_id == "1":
                return f"{self.employee} found {self.battery_desc}.\nNow {self.employee} is {self.emotion}, feeling that the task was {self.adjective}."
            elif self.work_type_id == "2":
                return (f"{self.employee} received {self.battery_desc} from {self.client_status} {self.client}.\n"
                        f"State: {self.battery_state}\n"
                        f"Actions: {self.battery_actions}\n"
                        f"Now {self.employee} is {self.emotion}, feeling that the task was {self.adjective}.")
            elif self.work_type_id == "3":
                return f"{self.employee} shipped {self.battery_desc} to {self.client_status} {self.client}.\nNow {self.employee} is {self.emotion}, feeling that the task was {self.adjective}."
            elif self.work_type_id == "4":
                return f"{self.employee} moved {self.battery_desc} from {self.old_location} to {self.new_location}.\nNow {self.employee} is {self.emotion}, feeling that the task was {self.adjective}."
            elif self.work_type_id == "20":
                return f"{self.employee} took a picture of {self.battery_desc}.\nNow {self.employee} is {self.emotion}, feeling that the task was {self.adjective}."
            else:
                return "No task description available."

    def fetch_location_desc(self, location_id):
        """Get location description from database."""
        if location_id:
            self.rds_cursor.execute("SELECT location_description FROM locations WHERE location_id = %s", (location_id,))
            result = self.rds_cursor.fetchall()
            return result[0][0] if result else "Unknown Location"
        return "Unknown Location"

    def fetch_client_status_desc(self, client_status_id):
        """Get client status description."""
        self.rds_cursor.execute("SELECT client_status_desc FROM client_status WHERE client_status_id = %s", (client_status_id,))
        result = self.rds_cursor.fetchall()
        return result[0][0] if result else "Unknown Status"

    def submit_request(self):
        """You can fill your own database submission here later."""
        pass

    def submit_report(self):
        """Save report into Excel."""
        self.request_cell = self.report_sheet[f"A{self.request_id + 1}"]
        self.timestamp_cell = self.report_sheet[f"B{self.request_id + 1}"]
        self.report_cell = self.report_sheet[f"C{self.request_id + 1}"]

        self.report_cell.alignment = Alignment(wrap_text=True)
        self.request_cell.value = self.request_id
        self.timestamp_cell.value = self.request_timestamp.get()
        self.report_cell.value = self.report_text.get("1.0", tk.END)

        self.adjust_cell_width()

    def adjust_cell_width(self):
        """Auto-size columns based on content length."""
        for col in self.report_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for row in range(1, self.report_sheet.max_row + 1):
                cell = self.report_sheet[f"{column}{row}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            self.report_sheet.column_dimensions[column].width = adjusted_width

    def reset_report(self):
        """Clear all variables."""
        self.request_id = None
        self.curr_time = None
        self.request_timestamp = tk.StringVar()
        self.serial_num = None
        self.work_type_id = None
        self.user_id = None
        self.state_id = "1"
        self.client_id = "1"
        self.actions = ["1"]
        self.old_location_id = None
        self.new_location_id = None
        self.picture = None
        self.part_num = None
        self.item_type = None
        self.emotion = None
        self.adjective = None
        self.report = tk.StringVar()
